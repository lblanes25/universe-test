"""Excel output helpers with consistent header formatting and conditional fills."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)

FILL_RED = PatternFill("solid", fgColor="F8CBAD")
FILL_YELLOW = PatternFill("solid", fgColor="FFE699")
FILL_GRAY = PatternFill("solid", fgColor="D9D9D9")
FILL_GREEN = PatternFill("solid", fgColor="C6E0B4")
FILL_RED_STRONG = PatternFill("solid", fgColor="F4B084")


def _format_sheet(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 10
        for cell in col_cells:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
    ws.freeze_panes = "A2"


def _apply_flag_fills(ws, priority_col: str | None) -> None:
    if priority_col is None:
        return
    headers = [c.value for c in ws[1]]
    if priority_col not in headers:
        return
    idx = headers.index(priority_col) + 1
    for row in ws.iter_rows(min_row=2):
        val = row[idx - 1].value
        fill = None
        if val == "HIGH":
            fill = FILL_RED
        elif val == "MEDIUM":
            fill = FILL_YELLOW
        elif val == "LOW":
            fill = FILL_GRAY
        if fill:
            for cell in row:
                cell.fill = fill


def _apply_coverage_fills(ws) -> None:
    headers = [c.value for c in ws[1]]
    if "In Scope" in headers:
        idx = headers.index("In Scope") + 1
        for row in ws.iter_rows(min_row=2):
            if row[idx - 1].value is True or row[idx - 1].value == "Yes":
                row[idx - 1].fill = FILL_GREEN
    if "Overdue Flag" in headers:
        idx = headers.index("Overdue Flag") + 1
        for row in ws.iter_rows(min_row=2):
            if row[idx - 1].value == "Yes":
                row[idx - 1].fill = FILL_RED_STRONG


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            safe = df if not df.empty else pd.DataFrame({"(empty)": []})
            safe.to_excel(writer, sheet_name=name[:31], index=False)
        wb = writer.book
        for ws in wb.worksheets:
            _format_sheet(ws)
            _apply_flag_fills(ws, "Priority")
            _apply_coverage_fills(ws)
